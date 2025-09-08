import streamlit as st
import numpy as np
import pandas as pd
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference

# Conversion table from point rankings to unit day values ($/user day)
POINT_VALUE_TABLE = pd.DataFrame(
    {
        "Points": [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100],
        "General Recreation": [
            4.87,
            5.78,
            6.39,
            7.31,
            9.13,
            10.35,
            11.26,
            11.87,
            13.09,
            14.00,
            14.61,
        ],
        "General Fishing and Hunting": [
            7.00,
            7.91,
            8.52,
            9.44,
            10.35,
            11.26,
            12.48,
            13.09,
            14.00,
            14.31,
            14.61,
        ],
        "Specialized Fishing and Hunting": [
            34.09,
            35.01,
            35.62,
            36.53,
            37.44,
            41.10,
            44.75,
            47.49,
            51.14,
            54.80,
            57.84,
        ],
        "Specialized Recreation": [
            19.79,
            21.00,
            22.53,
            24.35,
            25.88,
            29.22,
            32.27,
            38.97,
            45.36,
            51.75,
            57.84,
        ],
    }
)


st.title("Economic Toolbox")


def ead_trapezoidal(prob, damages):
    """Return expected annual damage via trapezoidal integration."""
    prob = np.asarray(prob, dtype=float)
    damages = np.asarray(damages, dtype=float)
    return float(
        np.sum(0.5 * (damages[:-1] + damages[1:]) * (prob[:-1] - prob[1:]))
    )


def updated_storage_cost(tc, sp, storage_reallocated, total_usable_storage):
    """Compute updated cost of storage for reservoir reallocations."""
    tc = float(tc)
    sp = float(sp)
    storage_reallocated = float(storage_reallocated)
    total_usable_storage = float(total_usable_storage)
    return (tc - sp) * storage_reallocated / total_usable_storage


def interest_during_construction(
    total_initial_cost,
    rate,
    months,
    *,
    costs=None,
    timings=None,
    normalize=True,
):
    """Compute interest during construction (IDC).

    Parameters
    ----------
    total_initial_cost : float
        Total initial cost excluding IDC. Used when ``costs`` is not provided.
    rate : float
        Annual interest rate expressed as a decimal (e.g., ``0.05`` for 5%).
    months : int
        Construction period in months.
    costs : list[float], optional
        Explicit costs for each month. If provided, ``normalize`` is ignored.
    timings : list[str], optional
        Timing of each cost within the month: ``"beginning"``, ``"middle"``, or
        ``"end"``. Defaults to ``"middle"`` for any unspecified timing.
    normalize : bool, default ``True``
        When ``True`` and ``costs`` is ``None``, the ``total_initial_cost`` is
        distributed evenly across all months with the first month treated as a
        beginning-of-month expenditure and remaining months at midpoints.

    Returns
    -------
    float
        Interest accrued during construction.
    """
    if months <= 0:
        return 0.0

    monthly_rate = rate / 12.0

    if costs is None:
        if not normalize:
            # Legacy approximation assuming evenly spread costs.
            years = months / 12.0
            return total_initial_cost * rate * years / 8

        monthly_cost = total_initial_cost / months
        costs = [monthly_cost] * months
        timings = ["beginning"] + ["middle"] * (months - 1)
    else:
        if timings is None:
            timings = ["middle"] * len(costs)

    idc = 0.0
    for i, cost in enumerate(costs, start=1):
        timing = timings[i - 1]
        if timing == "beginning":
            remaining = months - i + 1
        elif timing == "end":
            remaining = months - i
        else:  # middle
            remaining = months - i + 0.5
        idc += cost * monthly_rate * remaining
    return idc


def capital_recovery_factor(rate, periods):
    """Return capital recovery factor for a given discount rate and period."""
    if rate == 0:
        return 1 / periods
    return rate * (1 + rate) ** periods / ((1 + rate) ** periods - 1)


def build_excel():
    """Assemble all inputs, results, and README into an Excel workbook."""
    buffer = BytesIO()
    wb = Workbook()

    # EAD inputs sheet
    ws_ead_inputs = wb.active
    ws_ead_inputs.title = "EAD Inputs"
    table = st.session_state.get("table")
    if isinstance(table, pd.DataFrame):
        for row in dataframe_to_rows(table, index=False, header=True):
            ws_ead_inputs.append(row)

    # Charts sheet
    charts_for_export = st.session_state.get("charts_for_export", [])
    if charts_for_export:
        ws_charts = wb.create_sheet("Charts")
        for chart_info in charts_for_export:
            df_chart = chart_info["data"]
            start_row = ws_charts.max_row + 2 if ws_charts.max_row > 1 else 1
            for row in dataframe_to_rows(df_chart, index=False, header=True):
                ws_charts.append(row)
            end_row = start_row + len(df_chart)
            chart = LineChart()
            chart.title = chart_info["title"]
            chart.y_axis.title = df_chart.columns[1]
            chart.x_axis.title = df_chart.columns[0]
            data_ref = Reference(
                ws_charts, min_col=2, min_row=start_row, max_row=end_row
            )
            chart.add_data(data_ref, titles_from_data=True)
            cats_ref = Reference(
                ws_charts, min_col=1, min_row=start_row + 1, max_row=end_row
            )
            chart.set_categories(cats_ref)
            ws_charts.add_chart(chart, f"E{start_row}")

    # EAD results
    if st.session_state.get("ead_results"):
        ws_ead = wb.create_sheet("EAD Results")
        for k, v in st.session_state.ead_results.items():
            ws_ead.append([k, v])
        diffs = st.session_state.get("ead_differences", {})
        pct_changes = st.session_state.get("ead_percent_changes", {})
        if diffs:
            ws_ead.append([])
            for k, v in diffs.items():
                ws_ead.append([f"{k} - Damage 1", v])
        if pct_changes:
            ws_ead.append([])
            for k, v in pct_changes.items():
                ws_ead.append([f"{k} % change from Damage 1", v])

    # Updated storage inputs and result
    if st.session_state.get("storage_inputs") or st.session_state.get("storage_cost"):
        ws_storage = wb.create_sheet("Updated Storage")
        for k, v in st.session_state.get("storage_inputs", {}).items():
            ws_storage.append([k, v])
        if "storage_cost" in st.session_state:
            ws_storage.append(["Updated cost of storage", st.session_state.storage_cost])

    # Annualizer inputs, future costs, and summary
    if (
        st.session_state.get("annualizer_inputs")
        or st.session_state.get("future_costs_df") is not None
        or st.session_state.get("annualizer_summary")
    ):
        ws_ann = wb.create_sheet("Annualizer")
        for k, v in st.session_state.get("annualizer_inputs", {}).items():
            ws_ann.append([k, v])
        if st.session_state.get("annualizer_inputs"):
            ws_ann.append([])
        future_df = st.session_state.get("future_costs_df", pd.DataFrame())
        if not future_df.empty:
            for row in dataframe_to_rows(future_df, index=False, header=True):
                ws_ann.append(row)
            ws_ann.append([])
        for k, v in st.session_state.get("annualizer_summary", {}).items():
            ws_ann.append([k, v])

    # UDV analysis inputs and result
    if (
        st.session_state.get("udv_inputs")
        or st.session_state.get("udv_benefit")
    ):
        ws_rec = wb.create_sheet("UDV Analysis")
        for k, v in st.session_state.get("udv_inputs", {}).items():
            ws_rec.append([k, v])
        if "udv_benefit" in st.session_state:
            ws_rec.append(["Annual Recreation Benefit", st.session_state.udv_benefit])

    # Water demand forecast results
    water_df = st.session_state.get("water_demand_results")
    if isinstance(water_df, pd.DataFrame) and not water_df.empty:
        ws_water = wb.create_sheet("Water Demand")
        for row in dataframe_to_rows(water_df, index=False, header=True):
            ws_water.append(row)

    # README sheet
    readme_lines = Path("README.md").read_text().splitlines()
    ws_readme = wb.create_sheet("README")
    for line in readme_lines:
        ws_readme.append([line])

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_button():
    """Render a download button for the current workbook."""
    buffer = build_excel()
    st.download_button(
        label="Export to Excel",
        data=buffer,
        file_name="econ_toolbox.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Export all inputs, results, and the README as an Excel file.",
    )


def ead_calculator():
    """EAD calculator with data entry, charts, and results."""
    st.write(
        "Compute expected annual damages using the U.S. Army Corps of Engineers trapezoidal method."
    )
    st.caption(
        "Reference: U.S. Army Corps of Engineers, Engineering Manual 1110-2-1619 (1996)."
    )

    st.subheader("Inputs")
    st.info(
        "Fill in the frequency and damage values below. Use the checkbox to add a stage column and the button to insert additional damage columns as needed."
    )

    if "num_damage_cols" not in st.session_state:
        st.session_state.num_damage_cols = 1

    if "table" not in st.session_state:
        st.session_state.table = pd.DataFrame(
            {
                "Frequency": [1.00, 0.50, 0.20, 0.10, 0.04, 0.02, 0.01, 0.005, 0.002],
                "Damage 1": [
                    0,
                    10000,
                    40000,
                    80000,
                    120000,
                    160000,
                    200000,
                    250000,
                    300000,
                ],
            }
        )

    include_stage = st.checkbox(
        "Include stage column",
        value="Stage" in st.session_state.table.columns,
        help="Add a column for stage values to enable stage-related charts.",
    )
    if include_stage and "Stage" not in st.session_state.table.columns:
        st.session_state.table.insert(1, "Stage", [None] * len(st.session_state.table))
    elif not include_stage and "Stage" in st.session_state.table.columns:
        st.session_state.table.drop(columns="Stage", inplace=True)

    if st.button(
        "Add damage column",
        help="Insert another damage column to compare scenarios.",
    ):
        st.session_state.num_damage_cols += 1
        st.session_state.table[f"Damage {st.session_state.num_damage_cols}"] = [
            None
        ] * len(st.session_state.table)

    column_config = {
        "Frequency": st.column_config.NumberColumn(
            "Frequency", min_value=0, max_value=1, step=0.001
        )
    }
    for col in st.session_state.table.columns:
        if col.startswith("Damage"):
            column_config[col] = st.column_config.NumberColumn(
                col, min_value=0, step=1000, format="$%d"
            )

    with st.form("data_table_form"):
        data = st.data_editor(
            st.session_state.table,
            num_rows="dynamic",
            use_container_width=True,
            key="table_editor",
            column_config=column_config,
        )
        submitted = st.form_submit_button(
            "Save table", help="Apply edits to the table above."
        )
    if submitted:
        st.session_state.table = data

    damage_cols = [c for c in st.session_state.table.columns if c.startswith("Damage")]
    charts_for_export = []
    chart_data = (
        st.session_state.table.dropna(subset=["Frequency"])
        .sort_values("Frequency")
        .set_index("Frequency")[damage_cols]
    )
    if not chart_data.empty and damage_cols:
        st.subheader("Damage-Frequency Curve")
        selected_damage = st.selectbox(
            "Select damage column",
            damage_cols,
            key="df_damage",
            help="Choose which damage column to visualize.",
        )
        st.line_chart(chart_data[[selected_damage]])
        charts_for_export.append(
            {
                "title": "Damage-Frequency Curve",
                "data": chart_data[[selected_damage]].reset_index(),
            }
        )

    if "Stage" in st.session_state.table.columns:
        stage_df = (
            st.session_state.table.dropna(subset=["Stage"])
            .assign(Stage=lambda df: pd.to_numeric(df["Stage"], errors="coerce"))
            .dropna(subset=["Stage"])
            .set_index("Stage")
        )
        if not stage_df.empty:
            if damage_cols:
                st.subheader("Stage-Damage Curve")
                dmg_col = st.selectbox(
                    "Select damage column (stage)",
                    damage_cols,
                    key="stage_damage",
                    help="Damage column to plot against stage values.",
                )
                st.line_chart(stage_df[[dmg_col]])
                charts_for_export.append(
                    {
                        "title": "Stage-Damage Curve",
                        "data": stage_df[[dmg_col]].reset_index(),
                    }
                )
            if "Frequency" in stage_df.columns:
                st.subheader("Stage-Frequency Curve")
                st.line_chart(stage_df["Frequency"])
                charts_for_export.append(
                    {
                        "title": "Stage-Frequency Curve",
                        "data": stage_df[["Frequency"]].reset_index(),
                    }
                )

    st.session_state.charts_for_export = charts_for_export

    st.subheader("EAD Results")
    st.info(
        "Click the button below to compute expected annual damages for each damage column using trapezoidal integration."
    )
    if st.button(
        "Calculate EAD",
        help="Run the trapezoidal method on the frequency and damage data.",
    ):
        df = st.session_state.table.dropna(subset=["Frequency"]).sort_values(
            "Frequency", ascending=False
        )
        freq = df["Frequency"].to_numpy()
        if not np.isclose(freq[0], 1.0) or np.any(np.diff(freq) > 0):
            st.warning("Frequencies should start at 1 and monotonically decrease to 0.")
        missing_zero = freq[-1] != 0
        if missing_zero:
            st.info(
                "Final frequency not 0; appending zero-frequency point using last damage value."
            )
        results = {}
        for col in df.columns:
            if col.startswith("Damage"):
                damages = df[col].fillna(0).to_numpy()
                freq_use = np.append(freq, 0.0) if missing_zero else freq
                damages_use = (
                    np.append(damages, damages[-1]) if missing_zero else damages
                )
                if len(freq_use) >= 2 and len(freq_use) == len(damages_use):
                    results[col] = ead_trapezoidal(freq_use, damages_use)
                else:
                    results[col] = None
        if not results:
            st.error("No damage columns found.")
        else:
            base_ead = results.get("Damage 1")
            differences = {}
            pct_changes = {}
            for col, val in results.items():
                if val is None:
                    st.error(
                        f"{col}: Ensure at least two paired frequency and damage values."
                    )
                else:
                    st.success(f"{col} Expected Annual Damage: ${val:,.2f}")
                    if col != "Damage 1" and base_ead is not None:
                        diff = val - base_ead
                        differences[col] = diff
                        pct = (diff / base_ead * 100) if base_ead != 0 else np.nan
                        pct_changes[col] = pct
                        sign = "+" if diff >= 0 else "-"
                        st.info(
                            f"Difference from Damage 1: {sign}${abs(diff):,.2f} ({pct:+.2f}%)"
                        )
            st.session_state.ead_results = results
            st.session_state.ead_differences = differences
            st.session_state.ead_percent_changes = pct_changes

    export_button()


def storage_calculator():
    """Updated cost of storage calculator."""
    st.header("Updated Cost of Storage Calculator")
    st.caption(
        "Reference: Civil Works Construction Cost Index System (CWCCIS) and Engineering News Record (ENR)."
    )
    st.info("Estimate the updated cost of storage for a reservoir reallocation.")

    with st.form("storage_form"):
        tc = st.number_input(
            "Total Joint Use Construction Cost (TC)",
            min_value=0.0,
            value=1000000.0,
            help="Total joint-use construction costs updated using CWCCIS and ENR.",
        )
        sp = st.number_input(
            "Water Supply Specific Costs (SP)",
            min_value=0.0,
            value=100000.0,
            help="Costs of identifiable project features for a specific purpose, updated using CWCCIS and ENR.",
        )
        storage_reallocated = st.number_input(
            "Estimated Storage to be Reallocated (ac-ft)",
            min_value=0.0,
            value=1000.0,
            help="Estimated volume of storage being reallocated.",
        )
        total_usable_storage = st.number_input(
            "Total usable storage space (ac-ft)",
            min_value=0.0,
            value=10000.0,
            help="Total usable storage capacity of the project.",
        )
        compute_storage = st.form_submit_button(
            "Compute updated cost",
            help="Calculate the updated cost of storage.",
        )

    if compute_storage:
        cost = updated_storage_cost(
            tc, sp, storage_reallocated, total_usable_storage
        )
        st.success(f"Updated cost of storage: ${cost:,.2f}")
        st.session_state.storage_cost = cost
        st.session_state.storage_inputs = {
            "Total Joint Use Construction Cost (TC)": tc,
            "Water Supply Specific Costs (SP)": sp,
            "Estimated Storage to be Reallocated (ac-ft)": storage_reallocated,
            "Total usable storage space (ac-ft)": total_usable_storage,
        }

    export_button()


def annualizer_calculator():
    """Project cost annualizer."""
    st.header("Project Cost Annualizer")
    st.info("Calculate annualized project costs and benefit-cost ratio.")

    if "num_future_costs" not in st.session_state:
        st.session_state.num_future_costs = 0

    if st.button("Add planned future cost"):
        st.session_state.num_future_costs += 1

    with st.form("annualizer_form"):
        first_cost = st.number_input(
            "Project First Cost ($)", min_value=0.0, value=0.0
        )
        real_estate_cost = st.number_input(
            "Real Estate Cost ($)", min_value=0.0, value=0.0
        )
        ped_cost = st.number_input("PED Cost ($)", min_value=0.0, value=0.0)
        monitoring_cost = st.number_input(
            "Monitoring Cost ($)", min_value=0.0, value=0.0
        )
        idc_rate = st.number_input(
            "Interest Rate (%) - For Interest During Construction",
            min_value=0.0,
            value=0.0,
        )
        construction_months = st.number_input(
            "Construction Period (Months)", min_value=0, value=0, step=1
        )
        idc_method = st.radio(
            "IDC Cost Distribution",
            [
                "Normalize over construction period",
                "Specify per-period costs",
            ],
            index=0,
        )
        period_costs = []
        period_timings = []
        if idc_method == "Specify per-period costs":
            for m in range(1, int(construction_months) + 1):
                c = st.number_input(
                    f"Cost in month {m} ($)",
                    min_value=0.0,
                    value=0.0,
                    key=f"month_cost_{m}",
                )
                t = st.selectbox(
                    f"Timing for month {m}",
                    ["beginning", "middle", "end"],
                    key=f"month_time_{m}",
                    index=1,
                )
                period_costs.append(c)
                period_timings.append(t)
        annual_om = st.number_input(
            "Annual O&M Cost ($)", min_value=0.0, value=0.0
        )
        annual_benefits = st.number_input(
            "Benefits (Annual, $)", min_value=0.0, value=0.0
        )
        base_year = st.number_input("Base Year (Year)", min_value=0, step=1, value=0)
        discount_rate = st.number_input(
            "Discount Rate (%)", min_value=0.0, value=0.0
        )
        period_analysis = st.number_input(
            "Period of Analysis (Years)", min_value=1, step=1, value=1
        )

        future_costs = []
        for i in range(st.session_state.num_future_costs):
            cost = st.number_input(
                f"Planned Future Cost {i + 1} ($)",
                min_value=0.0,
                value=0.0,
                key=f"fcost_{i}",
            )
            year = st.number_input(
                f"Year of Cost {i + 1}",
                min_value=0,
                step=1,
                value=st.session_state.get(f"fyear_{i}", base_year),
                key=f"fyear_{i}",
            )
            future_costs.append((cost, year))

        compute_annual = st.form_submit_button("Compute Annual Costs")

    if compute_annual:
        initial_cost = first_cost + real_estate_cost + ped_cost + monitoring_cost
        if idc_method == "Specify per-period costs":
            idc = interest_during_construction(
                initial_cost,
                idc_rate / 100.0,
                construction_months,
                costs=period_costs,
                timings=period_timings,
                normalize=False,
            )
        else:
            idc = interest_during_construction(
                initial_cost, idc_rate / 100.0, construction_months
            )
        total_initial = initial_cost + idc
        dr = discount_rate / 100.0
        future_details = []
        for cost, year in future_costs:
            pv_factor = 1 / ((1 + dr) ** (year - base_year))
            pv = cost * pv_factor
            future_details.append(
                {
                    "Cost": cost,
                    "Year": year,
                    "PV Factor": pv_factor,
                    "Present Value": pv,
                }
            )
        pv_future = sum(item["Present Value"] for item in future_details)
        total_investment = total_initial + pv_future
        crf = capital_recovery_factor(dr, period_analysis)
        annual_construction = total_investment * crf
        annual_total = annual_construction + annual_om
        bcr = annual_benefits / annual_total if annual_total else np.nan

        if future_details:
            future_df = pd.DataFrame(future_details)
            st.write("Planned Future Costs (Present Values)")
            st.table(future_df)
            st.success(f"Present Value of Future Costs: ${pv_future:,.2f}")
            st.session_state.future_costs_df = future_df
        else:
            st.session_state.future_costs_df = pd.DataFrame()

        st.success(f"Interest During Construction: ${idc:,.2f}")
        st.success(f"Total Cost/Investment: ${total_investment:,.2f}")
        st.success(f"Capital Recovery Factor: {crf:.4f}")
        st.success(
            f"Annual Construction Cost including O&M: ${annual_total:,.2f}"
        )
        st.success(f"Benefit-Cost Ratio: {bcr:,.2f}")

        st.session_state.annualizer_summary = {
            "Interest During Construction": idc,
            "Total Cost/Investment": total_investment,
            "Capital Recovery Factor": crf,
            "Annual Cost including O&M": annual_total,
            "Benefit-Cost Ratio": bcr,
        }
        st.session_state.annualizer_inputs = {
            "Project First Cost ($)": first_cost,
            "Real Estate Cost ($)": real_estate_cost,
            "PED Cost ($)": ped_cost,
            "Monitoring Cost ($)": monitoring_cost,
            "Interest Rate (%)": idc_rate,
            "Construction Period (Months)": construction_months,
            "IDC Cost Distribution": idc_method,
            "Annual O&M Cost ($)": annual_om,
            "Benefits (Annual, $)": annual_benefits,
            "Base Year": base_year,
            "Discount Rate (%)": discount_rate,
            "Period of Analysis (Years)": period_analysis,
        }

    export_button()


def udv_analysis():
    """Unit Day Value recreation benefit calculator."""
    st.header("Recreation Benefit (Unit Day Value)")
    st.info(
        "Estimate annual recreation benefits using USACE Unit Day Values (UDV).",
    )
    tab_calc, tab_rank = st.tabs(["Calculator", "Ranking Criteria"])
    with tab_calc:
        rec_type = st.selectbox(
            "Recreation Type",
            ["General", "Specialized"],
            help="Select the type of recreation experience.",
        )
        if rec_type == "General":
            activity = st.selectbox(
                "General Activity Type",
                ["General Recreation", "Fishing and Hunting"],
                help="Select the general recreation category.",
            )
        else:
            activity = st.selectbox(
                "Specialized Activity Type",
                ["Fishing and Hunting", "Other (e.g., Boating)"],
                help="Select the specialized recreation category.",
            )
        points = st.number_input(
            "Point Value",
            min_value=0.0,
            max_value=100.0,
            value=0.0,
            step=1.0,
            help="Total recreation ranking points (0-100).",
        )
        column_map = {
            ("General", "General Recreation"): "General Recreation",
            ("General", "Fishing and Hunting"): "General Fishing and Hunting",
            ("Specialized", "Fishing and Hunting"): "Specialized Fishing and Hunting",
            ("Specialized", "Other (e.g., Boating)"): "Specialized Recreation",
        }
        table_col = column_map[(rec_type, activity)]
        udv_calc = float(
            np.interp(
                points, POINT_VALUE_TABLE["Points"], POINT_VALUE_TABLE[table_col]
            )
        )
        udv_value = st.number_input(
            "Unit Day Value ($/user day)",
            min_value=0.0,
            value=udv_calc,
            help="Override if updated UDV schedules are available.",
        )
        user_days = st.number_input(
            "Expected Annual User Days",
            min_value=0.0,
            value=0.0,
            step=1.0,
        )
        visitation = st.number_input(
            "Expected Visitation",
            min_value=0.0,
            value=1.0,
            step=1.0,
            help="Multiplier applied to the expected annual user days.",
        )
        if st.button("Compute Recreation Benefit"):
            total_user_days = user_days * visitation
            benefit = udv_value * total_user_days
            st.success(f"Annual Recreation Benefit: ${benefit:,.2f}")
            st.info(f"Adjusted Annual User Days: {total_user_days:,.2f}")
            st.session_state.udv_benefit = benefit
            st.session_state.udv_inputs = {
                "Recreation Type": rec_type,
                "Activity Type": activity,
                "Point Value": points,
                "Unit Day Value": udv_value,
                "Expected Annual User Days": user_days,
                "Expected Visitation": visitation,
                "Adjusted Annual User Days": total_user_days,
            }
    with tab_rank:
        st.subheader(
            "Table 1. Guidelines for Assigning Points for General Recreation"
        )
        criteria_table = {
            "Criteria": [
                "Recreation experience",
                "Availability of opportunity",
                "Carrying capacity",
                "Accessibility",
                "Environmental quality",
            ],
            "Very Low": [
                "Two general activities (0-4)",
                "Several within 1 hr travel time; a few within 30 min (0-3)",
                "Minimum facility for public health and safety (0-2)",
                "Limited access by any means to site or within site (0-3)",
                "Low aesthetic quality; factors significantly lower quality (0-2)",
            ],
            "Low": [
                "Several general activities (5-10)",
                "Several within 1 hr travel time; none within 30 min (4-6)",
                "Basic facility to conduct activity(ies) (3-5)",
                "Fair access, poor quality roads to site; limited access within site (4-6)",
                "Average aesthetic quality; factors exist that lower quality (3-6)",
            ],
            "Moderate": [
                "Several general activities; one high quality value activity (11-16)",
                "One or two within 1 hr travel time; none within 30 min (7-10)",
                "Adequate facilities to conduct activity without resource deterioration (6-8)",
                "Fair access, fair road to site; fair access, good roads within site (7-10)",
                "Above average aesthetic quality; limiting factors can be rectified (7-10)",
            ],
            "High": [
                "Several general activities; more than one high quality value activity (17-23)",
                "None within 1 hr travel time; one or two within 2 hr travel time (11-14)",
                "Optimum facilities to conduct activity at site (9-11)",
                "Good access, good roads to site; fair access, good roads within site (11-14)",
                "High aesthetic quality; no factors exist that lower quality (11-15)",
            ],
            "Very High": [
                "Numerous high quality activities (24-30)",
                "None within 2 hr travel time (15-18)",
                "Ultimate potential facilities to achieve intent of selected alternative (12-14)",
                "Good access, high standard road to site; good access within site (15-18)",
                "Outstanding aesthetic quality; no factors exist that lower quality (16-20)",
            ],
        }
        st.table(pd.DataFrame(criteria_table).set_index("Criteria"))
        st.subheader("Table 2. Conversion of Points to Dollar Values")
        st.table(POINT_VALUE_TABLE.set_index("Points"))
    export_button()


def water_demand_forecast():
    """Municipal and industrial water demand forecast."""
    st.header("Water Demand Forecast")
    st.info(
        "Project municipal and industrial water demand using USACE guidance." \
        " Calculations follow ER 1105-2-100 methodology."
    )

    with st.form("water_demand_form"):
        base_year = st.number_input(
            "Base Year",
            min_value=0,
            value=2024,
            step=1,
            help="Starting year for the forecast.",
        )
        projection_years = st.number_input(
            "Projection Years",
            min_value=1,
            value=20,
            step=1,
            help="Number of years to project beyond the base year.",
        )
        base_pop = st.number_input(
            "Base Population",
            min_value=0.0,
            value=0.0,
            step=100.0,
            help="Population in the base year.",
        )
        growth_rate = (
            st.number_input(
                "Annual Population Growth Rate (%)",
                value=1.0,
                step=0.1,
                help="Average annual growth rate for population.",
            )
            / 100.0
        )
        per_capita = st.number_input(
            "Per-capita Municipal Demand (gallons/person/day)",
            min_value=0.0,
            value=100.0,
            help="Typical municipal use per person.",
        )
        industrial_factor = (
            st.number_input(
                "Industrial Demand Factor (% of municipal)",
                min_value=0.0,
                value=20.0,
                step=1.0,
                help="Industrial demand as a percent of municipal demand.",
            )
            / 100.0
        )
        system_losses = (
            st.number_input(
                "System Losses (%)",
                min_value=0.0,
                value=10.0,
                step=1.0,
                help="Distribution losses as a percent of total demand.",
            )
            / 100.0
        )
        submitted = st.form_submit_button("Run Forecast")

    if submitted:
        years = np.arange(base_year, base_year + projection_years + 1)
        pops = base_pop * (1 + growth_rate) ** np.arange(0, projection_years + 1)
        input_df = pd.DataFrame({"Year": years, "Population": pops})

        municipal_mgy = pops * per_capita * 365 / 1e6
        industrial_mgy = municipal_mgy * industrial_factor
        total_mgy = (municipal_mgy + industrial_mgy) * (1 + system_losses)
        result_df = pd.DataFrame(
            {
                "Year": years,
                "Population": pops.round(0).astype(int),
                "Municipal Demand (MGY)": municipal_mgy,
                "Industrial Demand (MGY)": industrial_mgy,
                "Total Demand (MGY)": total_mgy,
            }
        )
        st.session_state.water_input_table = input_df
        st.session_state.water_demand_results = result_df

    if st.session_state.get("water_input_table") is not None:
        st.subheader("Population Projections")
        st.table(st.session_state.water_input_table)

    result_df = st.session_state.get("water_demand_results", pd.DataFrame())
    if not result_df.empty:
        st.subheader("Forecast Results")
        st.table(result_df)
        st.line_chart(result_df.set_index("Year")["Total Demand (MGY)"])

    export_button()


def readme_page():
    """Display repository README."""
    st.header("ReadMe")
    st.markdown(Path("README.md").read_text())


section = st.sidebar.radio(
    "Navigate",
    [
        "EAD Calculator",
        "Updated Storage Cost",
        "Project Annualizer",
        "UDV Analysis",
        "Water Demand Forecast",
        "ReadMe",
    ],
)

if section == "EAD Calculator":
    ead_calculator()
elif section == "Updated Storage Cost":
    storage_calculator()
elif section == "Project Annualizer":
    annualizer_calculator()
elif section == "UDV Analysis":
    udv_analysis()
elif section == "Water Demand Forecast":
    water_demand_forecast()
else:
    readme_page()

